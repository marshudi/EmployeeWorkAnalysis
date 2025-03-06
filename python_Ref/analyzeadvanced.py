import pandas as pd
import argparse
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configuration
HIGHLIGHT_WORDS = {
    'Scrum Master': 'FF0000',  # Red
    'Project support': '0000FF',  # Blue
    'Deployment': '00FF00',  # Green
    'Testing': 'FFA500',  # Orange
    'Sanity': 'FF0000'  # Red
}

STATUS_FILLS = {
    'Critical': PatternFill(start_color='FF0000', fill_type='solid'),
    'Moderate': PatternFill(start_color='FFA500', fill_type='solid'),
    'Normal': PatternFill(start_color='00FF00', fill_type='solid')
}

ROLES_TO_REMOVE = ['IT Engineer']

def extract_summary(description):
    """Extract important keywords from description for summary"""
    if not isinstance(description, str):
        return ''
    
    cleaned = re.sub(r'\b[A-Z]+-\d+\b', '', description)
    cleaned = re.sub(r'[^a-zA-Z, ]', '', cleaned)
    
    found_keywords = []
    for phrase in HIGHLIGHT_WORDS:
        if re.search(rf'\b{re.escape(phrase)}\b', description, re.IGNORECASE):
            found_keywords.append(phrase)
    
    words = [word for word in cleaned.split() if len(word) > 4]
    found_keywords.extend([w for w in words if w.lower() in [
        'management', 'support', 'testing', 'deployment', 'planning'
    ]])
    
    return ', '.join(sorted(set(found_keywords), key=len, reverse=True))

def read_input_file(input_file):
    """Read and validate input file"""
    file_ext = os.path.splitext(input_file)[1].lower()
    
    if file_ext == '.csv':
        df = pd.read_csv(input_file)
    elif file_ext in ['.xls', '.xlsx']:
        df = pd.read_excel(input_file)
    else:
        raise ValueError("Unsupported file format. Use CSV or Excel.")

    df.columns = df.columns.str.strip().str.lower()
    
    column_mapping = {
        'emplyee name': 'employee_name',
        'employee name': 'employee_name',
        'description of activity': 'description_of_activity',
        'date': 'date',
        'hours': 'hours',
        'onsite/offsite': 'location',
        'employee role': 'role',
        'project name': 'project'
    }
    
    df.rename(columns=lambda x: column_mapping.get(x.strip().lower(), x), inplace=True)
    
    required_columns = {'date', 'employee_name', 'hours', 'description_of_activity',
                       'location', 'role', 'project'}
    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    return df

def clean_role(role_entry):
    """Clean and filter roles"""
    if pd.isna(role_entry):
        return ''
    
    # Split and clean roles
    roles = [r.strip() for r in re.split(r'[,/]', str(role_entry)) if r.strip()]
    
    # Remove unwanted roles (case-insensitive)
    filtered_roles = [
        r for r in roles 
        if not any(remove_role.lower() == r.lower() for remove_role in ROLES_TO_REMOVE)
    ]
    
    return ', '.join(sorted(set(filtered_roles))) if filtered_roles else ''

def analyze_employee_data(input_file):
    df = read_input_file(input_file)

    # Clean data
    df['location'] = df['location'].str.replace('-', ' ').str.title()
    df['role'] = df['role'].apply(clean_role)
    df['project'] = df['project'].str.replace(r'[^a-zA-Z0-9 ]', '', regex=True)

    # Group and aggregate
    aggregation = {
        'hours': 'sum',
        'date': 'count',
        'description_of_activity': lambda x: ' | '.join(set(x.dropna().astype(str))),
        'location': lambda x: ', '.join(sorted(set(x.dropna().astype(str)))),
        'role': lambda x: ', '.join(sorted(set(x.dropna().astype(str)))),
        'project': lambda x: ', '.join(sorted(set(x.dropna().astype(str))))
    }

    df_grouped = df.groupby('employee_name', as_index=False).agg(aggregation)

    # Process grouped data
    df_grouped['summary'] = df_grouped['description_of_activity'].apply(extract_summary)
    df_grouped = df_grouped.rename(columns={
        'date': 'Total Days',
        'employee_name': 'Employee Name',
        'hours': 'Total Hours'
    })

    # Calculate metrics
    df_grouped['Total Hours'] = pd.to_numeric(df_grouped['Total Hours'], errors='coerce').fillna(0)
    df_grouped['Man Days'] = (df_grouped['Total Hours'] / 8).round(2)
    df_grouped['Avg Hours per Day'] = (df_grouped['Total Hours'] / df_grouped['Total Days']).round(2)
    
    df_grouped['Status'] = df_grouped['Avg Hours per Day'].apply(
        lambda x: 'Critical' if x > 12 else 'Moderate' if x > 8 else 'Normal'
    )

    # Hide non-critical descriptions
    df_grouped['description_of_activity'] = df_grouped.apply(
        lambda row: row['description_of_activity'] if row['Status'] == 'Critical' else '', axis=1
    )

    # Final column order
    column_order = [
        'Man Days', 'Employee Name', 'location', 'role', 'project',
        'Total Hours', 'Avg Hours per Day', 'Status', 'summary', 'description_of_activity'
    ]
    df_grouped = df_grouped[column_order]

    # Save to Excel
    output_file = f"{os.path.splitext(input_file)[0]}_aggregated.xlsx"
    df_grouped.to_excel(output_file, index=False, sheet_name='Summary')

    # Apply formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Status column formatting
    status_col_idx = column_order.index('Status') + 1
    for row_num, status in enumerate(df_grouped['Status'], start=2):
        ws.cell(row=row_num, column=status_col_idx).fill = STATUS_FILLS[status]

    # Auto-adjust columns
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in ws[col[0].column_letter]), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(output_file)
    print(f"Aggregated analysis saved to {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Employee work analysis with role filtering")
    parser.add_argument("input_file", help="Path to input CSV/Excel file")
    args = parser.parse_args()
    
    analyze_employee_data(args.input_file)